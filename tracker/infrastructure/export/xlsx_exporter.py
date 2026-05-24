from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass(frozen=True)
class TrendTable:
    title: str
    headers: list[str]
    rows: list[list[str]]


@dataclass(frozen=True)
class PortfolioValueTrendData:
    report_quarter: str
    previous_quarter: str
    selected_managers: int
    analyzed_managers: int
    missing_current: int
    missing_previous: int
    growing_managers: int
    holding_managers: int
    reducing_managers: int
    previous_total_value_k: int
    current_total_value_k: int
    shares_analyzed_managers: int
    shares_growing_managers: int
    shares_holding_managers: int
    shares_reducing_managers: int
    previous_total_shares: int
    current_total_shares: int


@dataclass(frozen=True)
class TrendSummaryWorkbookData:
    report_quarter: str
    view_mode: str
    min_conf: float
    limit: int
    top_buy: TrendTable
    top_sell: TrendTable
    reversals: TrendTable | None
    portfolio_value_trend: PortfolioValueTrendData | None
    content_fingerprint: str
    call_options: TrendTable | None = None
    put_options: TrendTable | None = None


def _write_table_sheet(
    workbook: Workbook,
    sheet_title: str,
    table: TrendTable,
    start_row: int = 1,
) -> None:
    if sheet_title in workbook.sheetnames:
        sheet = workbook[sheet_title]
    else:
        sheet = workbook.create_sheet(title=sheet_title)

    row_idx = start_row

    # Table title
    sheet.cell(row=row_idx, column=1, value=table.title)
    sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    row_idx += 2

    if not table.rows:
        sheet.cell(row=row_idx, column=1, value="(empty)")
        row_idx += 1
        return

    # Headers
    for col_idx, header in enumerate(table.headers, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGNMENT_CENTER
    row_idx += 1

    # Rows
    for row in table.rows:
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _ALIGNMENT_LEFT
        row_idx += 1

    # Auto-width columns
    for col_idx, header in enumerate(table.headers, start=1):
        max_length = len(header)
        for row in table.rows:
            cell_value = row[col_idx - 1] if col_idx <= len(row) else ""
            max_length = max(max_length, len(str(cell_value)))
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = min(max_length + 4, 60)


def _format_value_k(value_k: int) -> str:
    value_billions = round(value_k / 1_000_000_000)
    return f"${value_billions:,}B"


def _format_int(value: int) -> str:
    return f"{value:,}"


def _format_signed_ratio(value: float) -> str:
    return f"{value * 100:+.1f}%"


def _format_ratio(value: float) -> str:
    return f"{value * 100:.1f}%"


def _portfolio_value_change_ratio(previous_value_k: int, current_value_k: int) -> float:
    if previous_value_k <= 0:
        return 0.0 if current_value_k <= 0 else 1.0
    return (current_value_k - previous_value_k) / previous_value_k


def _portfolio_value_direction(change_ratio: float) -> str:
    hold_band = 0.05
    if change_ratio > hold_band:
        return "Growing"
    if change_ratio < -hold_band:
        return "Reducing"
    return "Holding"


def _write_portfolio_value_trend_sheet(
    workbook: Workbook,
    data: PortfolioValueTrendData,
) -> None:
    sheet_title = "Portfolio Value Trend"
    if sheet_title in workbook.sheetnames:
        sheet = workbook[sheet_title]
    else:
        sheet = workbook.create_sheet(title=sheet_title)

    row_idx = 1

    # Title
    sheet.cell(row=row_idx, column=1, value="Hedge Funds Portfolio Value Trend (QoQ)")
    sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    row_idx += 2

    # Quarters and managers
    sheet.cell(row=row_idx, column=1, value="Compared quarters")
    sheet.cell(row=row_idx, column=2, value=f"{data.previous_quarter} -> {data.report_quarter}")
    row_idx += 1
    sheet.cell(row=row_idx, column=1, value="Managers analyzed")
    sheet.cell(row=row_idx, column=2, value=f"{data.analyzed_managers}/{data.selected_managers}")
    row_idx += 1
    if data.missing_current > 0:
        sheet.cell(row=row_idx, column=1, value="Missing current quarter")
        sheet.cell(row=row_idx, column=2, value=str(data.missing_current))
        row_idx += 1
    if data.missing_previous > 0:
        sheet.cell(row=row_idx, column=1, value="Missing previous quarter")
        sheet.cell(row=row_idx, column=2, value=str(data.missing_previous))
        row_idx += 1

    if data.analyzed_managers == 0:
        sheet.cell(row=row_idx, column=1, value="Not enough comparable snapshots to determine portfolio value direction.")
        row_idx += 1
        return

    row_idx += 1

    # Aggregate values
    aggregate_change_ratio = _portfolio_value_change_ratio(
        data.previous_total_value_k, data.current_total_value_k
    )
    aggregate_direction = _portfolio_value_direction(aggregate_change_ratio)
    sheet.cell(row=row_idx, column=1, value="Aggregate portfolio value")
    sheet.cell(
        row=row_idx,
        column=2,
        value=(
            f"{_format_value_k(data.previous_total_value_k)} -> "
            f"{_format_value_k(data.current_total_value_k)} "
            f"({_format_signed_ratio(aggregate_change_ratio)} {aggregate_direction})"
        ),
    )
    row_idx += 1

    if data.shares_analyzed_managers > 0:
        aggregate_shares_change_ratio = _portfolio_value_change_ratio(
            data.previous_total_shares, data.current_total_shares
        )
        aggregate_shares_direction = _portfolio_value_direction(aggregate_shares_change_ratio)
        sheet.cell(row=row_idx, column=1, value="Aggregate portfolio shares")
        sheet.cell(
            row=row_idx,
            column=2,
            value=(
                f"{_format_int(data.previous_total_shares)} -> "
                f"{_format_int(data.current_total_shares)} "
                f"({_format_signed_ratio(aggregate_shares_change_ratio)} {aggregate_shares_direction})"
            ),
        )
        row_idx += 1

    row_idx += 1

    # Value Direction Breakdown
    breakdown_headers = ["Direction", "Managers", "Share"]
    for col_idx, header in enumerate(breakdown_headers, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGNMENT_CENTER
    row_idx += 1

    value_rows = [
        ["Growing", str(data.growing_managers), _format_ratio(data.growing_managers / data.analyzed_managers)],
        ["Holding", str(data.holding_managers), _format_ratio(data.holding_managers / data.analyzed_managers)],
        ["Reducing", str(data.reducing_managers), _format_ratio(data.reducing_managers / data.analyzed_managers)],
    ]
    for row in value_rows:
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1

    if data.shares_analyzed_managers > 0:
        row_idx += 1
        if data.shares_analyzed_managers != data.analyzed_managers:
            sheet.cell(row=row_idx, column=1, value=f"Shares coverage: {data.shares_analyzed_managers}/{data.analyzed_managers}")
            row_idx += 1

        for col_idx, header in enumerate(breakdown_headers, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _ALIGNMENT_CENTER
        row_idx += 1

        shares_rows = [
            ["Growing", str(data.shares_growing_managers), _format_ratio(data.shares_growing_managers / data.shares_analyzed_managers)],
            ["Holding", str(data.shares_holding_managers), _format_ratio(data.shares_holding_managers / data.shares_analyzed_managers)],
            ["Reducing", str(data.shares_reducing_managers), _format_ratio(data.shares_reducing_managers / data.shares_analyzed_managers)],
        ]
        for row in shares_rows:
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    # Auto-width for first two columns
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 50
    sheet.column_dimensions["C"].width = 15


def _write_metadata_sheet(
    workbook: Workbook,
    data: TrendSummaryWorkbookData,
) -> None:
    sheet_title = "Metadata"
    if sheet_title in workbook.sheetnames:
        sheet = workbook[sheet_title]
    else:
        sheet = workbook.create_sheet(title=sheet_title)

    metadata_items: list[tuple[str, Any]] = [
        ("Report Quarter", data.report_quarter),
        ("View Mode", data.view_mode),
        ("Min Confidence", data.min_conf),
        ("Limit", data.limit),
        ("Content Fingerprint", data.content_fingerprint),
    ]

    for row_idx, (key, value) in enumerate(metadata_items, start=1):
        sheet.cell(row=row_idx, column=1, value=key).font = _HEADER_FONT
        sheet.cell(row=row_idx, column=2, value=value)

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 70


def write_trend_summary_workbook(path: Path, data: TrendSummaryWorkbookData) -> None:
    workbook = Workbook()

    # Remove default sheet; we'll create named ones
    workbook.remove(workbook.active)

    _write_table_sheet(workbook, "Top Buy Ideas", data.top_buy)
    _write_table_sheet(workbook, "Top Reduction Trends", data.top_sell)
    if data.reversals is not None:
        _write_table_sheet(workbook, "Reversals", data.reversals)
    if data.call_options is not None:
        _write_table_sheet(workbook, "Call Option Trends", data.call_options)
    if data.put_options is not None:
        _write_table_sheet(workbook, "Put Option Trends", data.put_options)
    if data.portfolio_value_trend is not None:
        _write_portfolio_value_trend_sheet(workbook, data.portfolio_value_trend)
    _write_metadata_sheet(workbook, data)

    workbook.save(str(path))


def read_content_fingerprint(path: Path) -> str | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        workbook = load_workbook(str(path))
        metadata_sheet = workbook["Metadata"]
        for row in metadata_sheet.iter_rows(min_row=1, max_row=metadata_sheet.max_row, values_only=False):
            if row[0].value == "Content Fingerprint" and len(row) > 1:
                return str(row[1].value) if row[1].value is not None else None
        return None
    except Exception:
        return None
