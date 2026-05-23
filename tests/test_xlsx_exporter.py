from __future__ import annotations

from pathlib import Path

import pytest
from tracker.infrastructure.export.xlsx_exporter import (
    PortfolioValueTrendData,
    TrendSummaryWorkbookData,
    TrendTable,
    read_content_fingerprint,
    write_trend_summary_workbook,
)


def _sample_portfolio_data() -> PortfolioValueTrendData:
    return PortfolioValueTrendData(
        report_quarter="2025Q4",
        previous_quarter="2025Q3",
        selected_managers=4,
        analyzed_managers=3,
        missing_current=0,
        missing_previous=1,
        growing_managers=1,
        holding_managers=1,
        reducing_managers=1,
        previous_total_value_k=600_000_000,
        current_total_value_k=577_000_000,
        shares_analyzed_managers=3,
        shares_growing_managers=1,
        shares_holding_managers=1,
        shares_reducing_managers=1,
        previous_total_shares=6_000,
        current_total_shares=5_750,
    )


def _sample_workbook_data() -> TrendSummaryWorkbookData:
    return TrendSummaryWorkbookData(
        report_quarter="2025Q4",
        view_mode="shortlist",
        min_conf=0.45,
        limit=8,
        top_buy=TrendTable(
            title="Top Buy Ideas",
            headers=["Instrument", "Setup", "Idea Score", "Support", "Confidence", "Freshness", "Top Contributors"],
            rows=[
                ["AAPL", "Strong", "0.8500", "4/1", "80%", "Fresh", "Fund A, Fund B"],
                ["MSFT", "Emerging", "0.7200", "3/2", "75%", "Fresh", "Fund C"],
            ],
        ),
        top_sell=TrendTable(
            title="Top Reduction Trends",
            headers=["Instrument", "Setup", "Idea Score", "Support", "Confidence", "Freshness", "Top Contributors"],
            rows=[
                ["TSLA", "Weakening", "-0.6500", "1/3", "70%", "Stale", "Fund D"],
            ],
        ),
        reversals=None,
        portfolio_value_trend=_sample_portfolio_data(),
        content_fingerprint="abc123def456",
    )


def test_write_trend_summary_workbook_creates_expected_sheets(tmp_path: Path) -> None:
    path = tmp_path / "trend_summary.xlsx"
    data = _sample_workbook_data()
    write_trend_summary_workbook(path, data)

    assert path.exists()

    from openpyxl import load_workbook

    workbook = load_workbook(str(path))
    sheet_names = workbook.sheetnames
    assert "Top Buy Ideas" in sheet_names
    assert "Top Reduction Trends" in sheet_names
    assert "Portfolio Value Trend" in sheet_names
    assert "Metadata" in sheet_names


def test_metadata_sheet_contains_fingerprint(tmp_path: Path) -> None:
    path = tmp_path / "trend_summary.xlsx"
    data = _sample_workbook_data()
    write_trend_summary_workbook(path, data)

    fingerprint = read_content_fingerprint(path)
    assert fingerprint == "abc123def456"


def test_read_content_fingerprint_returns_none_for_missing_file(tmp_path: Path) -> None:
    path = tmp_path / "nonexistent.xlsx"
    assert read_content_fingerprint(path) is None


def test_top_buy_sheet_has_correct_data(tmp_path: Path) -> None:
    path = tmp_path / "trend_summary.xlsx"
    data = _sample_workbook_data()
    write_trend_summary_workbook(path, data)

    from openpyxl import load_workbook

    workbook = load_workbook(str(path))
    sheet = workbook["Top Buy Ideas"]

    # Row 1: title, Row 2: blank, Row 3: headers, Row 4+: data
    assert sheet.cell(row=3, column=1).value == "Instrument"
    assert sheet.cell(row=3, column=2).value == "Setup"
    assert sheet.cell(row=4, column=1).value == "AAPL"
    assert sheet.cell(row=4, column=3).value == "0.8500"
    assert sheet.cell(row=5, column=1).value == "MSFT"


def test_reversals_sheet_created_when_present(tmp_path: Path) -> None:
    path = tmp_path / "trend_summary.xlsx"
    data = TrendSummaryWorkbookData(
        report_quarter="2025Q4",
        view_mode="raw",
        min_conf=0.45,
        limit=8,
        top_buy=TrendTable(title="Top Buy Trends", headers=["Ticker"], rows=[["AAPL"]]),
        top_sell=TrendTable(title="Top Sell Trends", headers=["Ticker"], rows=[["TSLA"]]),
        reversals=TrendTable(title="Reversals", headers=["Ticker"], rows=[["NVDA"]]),
        portfolio_value_trend=None,
        content_fingerprint="rev789",
    )
    write_trend_summary_workbook(path, data)

    from openpyxl import load_workbook

    workbook = load_workbook(str(path))
    assert "Reversals" in workbook.sheetnames
    sheet = workbook["Reversals"]
    assert sheet.cell(row=4, column=1).value == "NVDA"


def test_portfolio_value_trend_sheet_has_aggregate_data(tmp_path: Path) -> None:
    path = tmp_path / "trend_summary.xlsx"
    data = _sample_workbook_data()
    write_trend_summary_workbook(path, data)

    from openpyxl import load_workbook

    workbook = load_workbook(str(path))
    sheet = workbook["Portfolio Value Trend"]

    assert sheet.cell(row=1, column=1).value == "Hedge Funds Portfolio Value Trend (QoQ)"
    assert sheet.cell(row=3, column=1).value == "Compared quarters"
    assert sheet.cell(row=3, column=2).value == "2025Q3 -> 2025Q4"
