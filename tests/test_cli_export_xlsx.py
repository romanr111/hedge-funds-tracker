from __future__ import annotations

import importlib
import json
from datetime import datetime, timezone
from pathlib import Path

import pytest
from signals.domain.models import TrendStockSignal
from signals.infrastructure.storage.sqlite_state_repository import StateStore

cli_main_module = importlib.import_module("signals.interfaces.cli.main")


def _seed_trend_data(db_path: Path) -> None:
    now_iso = datetime.now(timezone.utc).isoformat()
    store = StateStore(db_path)
    try:
        store.replace_trend_stock_signals(
            "2025Q4",
            [
                TrendStockSignal(
                    report_quarter="2025Q4",
                    instrument_key="111111111",
                    cusip="111111111",
                    put_call=None,
                    issuer_name="Alpha Corp",
                    title="COM",
                    np_raw=0.12,
                    np_adj=0.12,
                    impulse_score=0.10,
                    accumulation_score=0.09,
                    confidence=0.80,
                    trend_ewma=0.09,
                    trend_delta=0.03,
                    breadth_buy_weight=0.20,
                    breadth_sell_weight=0.01,
                    buy_managers=4,
                    sell_managers=1,
                    crowding_hhi=0.20,
                    persistence_buy=2,
                    persistence_sell=0,
                    regime="STRONG_BUY",
                    contributors_json=json.dumps(
                        [
                            {"manager_name": "TCI Fund Management Ltd", "signal_value": 0.40},
                            {"manager_name": "Coatue Management LLC", "signal_value": 0.30},
                        ],
                        separators=(",", ":"),
                    ),
                    computed_at=now_iso,
                    freshness_multiplier=1.0,
                    freshness_ok=True,
                ),
            ],
        )
        # Seed snapshots for portfolio value trend computation
        for quarter, aum_value_k in (
            ("2025Q3", 600_000_000),
            ("2025Q4", 577_000_000),
        ):
            store.upsert_manager_quarter_snapshot(
                cik="0000000001",
                manager_name="Fund 1",
                report_quarter=quarter,
                report_date="2025-12-31",
                filing_date="2026-02-14",
                acceptance_datetime="2026-02-14T10:00:00Z",
                accession=f"0000000001-{quarter}",
                source_form="13F-HR",
                positions=[{"name": "Alpha Corp", "title": "COM", "cusip": "111111111", "value": 10_000, "shares": 1_000}],
                aum_value_k=aum_value_k,
            )
    finally:
        store.close()


def _seed_option_trend_data(db_path: Path) -> None:
    now_iso = datetime.now(timezone.utc).isoformat()
    store = StateStore(db_path)
    try:
        store.replace_trend_option_signals(
            "2025Q4",
            [
                TrendStockSignal(
                    report_quarter="2025Q4",
                    instrument_key="111111111|CALL",
                    cusip="111111111",
                    put_call="CALL",
                    issuer_name="Alpha Corp",
                    title="OPTION",
                    np_raw=0.08,
                    np_adj=0.08,
                    impulse_score=0.08,
                    accumulation_score=0.07,
                    confidence=0.80,
                    trend_ewma=0.07,
                    trend_delta=0.02,
                    breadth_buy_weight=0.20,
                    breadth_sell_weight=0.01,
                    buy_managers=3,
                    sell_managers=0,
                    crowding_hhi=0.20,
                    persistence_buy=2,
                    persistence_sell=0,
                    regime="STRONG_BUY",
                    contributors_json=json.dumps([{"manager_name": "Fund 1", "signal_value": 0.08}], separators=(",", ":")),
                    computed_at=now_iso,
                    freshness_multiplier=1.0,
                    freshness_ok=None,
                ),
            ],
        )
    finally:
        store.close()


def test_build_trend_summary_workbook_data_returns_expected_structure(tmp_path: Path) -> None:
    db_path = tmp_path / "signals.sqlite3"
    _seed_trend_data(db_path)
    store = StateStore(db_path)
    try:
        signals = store.list_trend_stock_signals("2025Q4")
        data = cli_main_module._build_trend_summary_workbook_data(
            store,
            "2025Q4",
            signals=signals,
            symbol_map={"111111111": "AAPL"},
            min_conf=0.45,
            limit=8,
            manager_ciks=["0000000001"],
            view="shortlist",
            show_reversals=False,
        )
    finally:
        store.close()

    assert data.report_quarter == "2025Q4"
    assert data.view_mode == "shortlist"
    assert data.top_buy.title == "Top Buy Ideas"
    assert len(data.top_buy.rows) == 1
    assert data.top_buy.rows[0][0] == "AAPL"
    assert data.top_sell.title == "Top Reduction Trends"
    assert data.portfolio_value_trend is not None
    assert data.content_fingerprint
    assert len(data.content_fingerprint) == 64  # SHA-256 hex


def test_build_trend_summary_workbook_data_includes_option_tables(tmp_path: Path) -> None:
    db_path = tmp_path / "signals.sqlite3"
    _seed_trend_data(db_path)
    _seed_option_trend_data(db_path)
    store = StateStore(db_path)
    try:
        signals = store.list_trend_stock_signals("2025Q4")
        data = cli_main_module._build_trend_summary_workbook_data(
            store,
            "2025Q4",
            signals=signals,
            symbol_map={"111111111": "AAPL"},
            min_conf=0.45,
            limit=8,
            manager_ciks=["0000000001"],
            view="shortlist",
            show_reversals=False,
        )
    finally:
        store.close()

    assert data.call_options is not None
    assert data.call_options.title == "Top Call Option Trends"
    assert data.call_options.rows[0][0] == "AAPL CALL"
    assert data.put_options is not None
    assert data.put_options.title == "Top Put Option Trends"
    assert data.content_fingerprint


def test_build_trend_summary_workbook_data_raw_view(tmp_path: Path) -> None:
    db_path = tmp_path / "signals.sqlite3"
    _seed_trend_data(db_path)
    store = StateStore(db_path)
    try:
        signals = store.list_trend_stock_signals("2025Q4")
        data = cli_main_module._build_trend_summary_workbook_data(
            store,
            "2025Q4",
            signals=signals,
            symbol_map={"111111111": "AAPL"},
            min_conf=0.45,
            limit=8,
            manager_ciks=["0000000001"],
            view="raw",
            show_reversals=False,
        )
    finally:
        store.close()

    assert data.view_mode == "raw"
    assert data.top_buy.title == "Top Buy Trends"
    assert data.top_sell.title == "Top Sell Trends"


def test_build_trend_summary_workbook_data_fingerprint_changes_with_view(tmp_path: Path) -> None:
    db_path = tmp_path / "signals.sqlite3"
    _seed_trend_data(db_path)
    store = StateStore(db_path)
    try:
        signals = store.list_trend_stock_signals("2025Q4")
        data_shortlist = cli_main_module._build_trend_summary_workbook_data(
            store,
            "2025Q4",
            signals=signals,
            symbol_map={"111111111": "AAPL"},
            min_conf=0.45,
            limit=8,
            manager_ciks=["0000000001"],
            view="shortlist",
            show_reversals=False,
        )
        data_raw = cli_main_module._build_trend_summary_workbook_data(
            store,
            "2025Q4",
            signals=signals,
            symbol_map={"111111111": "AAPL"},
            min_conf=0.45,
            limit=8,
            manager_ciks=["0000000001"],
            view="raw",
            show_reversals=False,
        )
    finally:
        store.close()

    assert data_shortlist.content_fingerprint != data_raw.content_fingerprint


def test_maybe_export_trend_summary_writes_file(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    db_path = tmp_path / "signals.sqlite3"
    _seed_trend_data(db_path)
    store = StateStore(db_path)
    export_path = tmp_path / "exports"
    logger = cli_main_module.logging.getLogger("test")
    try:
        cli_main_module._maybe_export_trend_summary(
            store,
            "2025Q4",
            export_xlsx_path=str(export_path),
            min_conf=0.45,
            limit=8,
            show_reversals=False,
            symbols_file="config/cusip_tickers.json",
            manager_ciks=["0000000001"],
            view="shortlist",
            dry_run=False,
            logger=logger,
        )
    finally:
        store.close()

    expected_file = export_path / "trend_summary_2025Q4.xlsx"
    assert expected_file.exists()


def test_maybe_export_trend_summary_writes_option_only_file(tmp_path: Path) -> None:
    db_path = tmp_path / "signals.sqlite3"
    _seed_option_trend_data(db_path)
    store = StateStore(db_path)
    export_path = tmp_path / "exports"
    logger = cli_main_module.logging.getLogger("test")
    try:
        cli_main_module._maybe_export_trend_summary(
            store,
            "2025Q4",
            export_xlsx_path=str(export_path),
            min_conf=0.45,
            limit=8,
            show_reversals=False,
            symbols_file="config/cusip_tickers.json",
            manager_ciks=[],
            view="shortlist",
            dry_run=False,
            logger=logger,
        )
    finally:
        store.close()

    expected_file = export_path / "trend_summary_2025Q4.xlsx"
    assert expected_file.exists()

    from openpyxl import load_workbook

    workbook = load_workbook(str(expected_file))
    assert "Call Option Trends" in workbook.sheetnames
    assert workbook["Call Option Trends"].cell(row=4, column=1).value.endswith("CALL")
